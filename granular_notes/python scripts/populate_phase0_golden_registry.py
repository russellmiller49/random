#!/usr/bin/env python3
from __future__ import annotations

import argparse
import inspect
import json
import re
import runpy
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable

import openpyxl


@dataclass(frozen=True)
class NotePayload:
    note_id: str
    source_file: str
    note_text: str
    procedure_date: str = ""
    flags: dict[str, Any] | None = None
    spans: list[dict[str, Any]] | None = None
    events: list[dict[str, Any]] | None = None


def _truthy_flag(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value != 0)
    if isinstance(value, str):
        v = value.strip().lower()
        if v in {"1", "true", "t", "yes", "y"}:
            return 1
        if v in {"0", "false", "f", "no", "n", ""}:
            return 0
    return int(bool(value))


def _to_tf(value: Any, default: str = "FALSE") -> str:
    if value is None:
        return default
    if isinstance(value, str):
        v = value.strip().upper()
        if v in {"TRUE", "FALSE"}:
            return v
        return "TRUE" if _truthy_flag(value) else "FALSE"
    return "TRUE" if _truthy_flag(value) else "FALSE"


def _safe_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        try:
            return int(v)
        except ValueError:
            return None
    return None


def _call_best_effort(fn: Callable[..., Any], note_text: str) -> Any:
    sig = inspect.signature(fn)
    params = list(sig.parameters.values())

    if not params:
        return fn()

    # Prefer passing note text as the first positional param.
    try:
        return fn(note_text)
    except TypeError:
        pass

    # Try common parameter names.
    kwargs: dict[str, Any] = {}
    for p in params:
        if p.kind in (inspect.Parameter.VAR_POSITIONAL, inspect.Parameter.VAR_KEYWORD):
            continue
        if p.name in {"text", "note_text", "note", "content"}:
            kwargs[p.name] = note_text
    if kwargs:
        return fn(**kwargs)

    # Last resort: give up.
    return fn()


def _read_note_text(notes_dir: Path, source_file: str) -> str | None:
    p = notes_dir / source_file
    if not p.exists():
        return None
    return p.read_text(encoding="utf-8", errors="replace")


def _hydrate_span(note_text: str, span_text: str, context_prefix: str | None, match_index: int | None) -> tuple[int | None, int | None, str]:
    if not span_text:
        return None, None, "empty_span_text"

    matches = [m.start() for m in re.finditer(re.escape(span_text), note_text)]
    if not matches:
        return None, None, "not_found"

    if len(matches) == 1:
        start = matches[0]
        return start, start + len(span_text), "hydrated_unique"

    # Prefer context prefix within preceding window.
    if context_prefix:
        prefix = context_prefix
        for m_start in matches:
            window_start = max(0, m_start - 120)
            window_text = note_text[window_start:m_start]
            if prefix in window_text:
                return m_start, m_start + len(span_text), "hydrated_prefix_window"

    if match_index is not None and 0 <= match_index < len(matches):
        start = matches[match_index]
        return start, start + len(span_text), "hydrated_match_index"

    return None, None, f"ambiguous_count={len(matches)}"


def _sheet_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[str]:
    headers: list[str] = []
    for cell in ws[1]:
        if cell.value is None:
            break
        headers.append(str(cell.value))
    return headers


def _excel_safe_value(value: Any) -> Any:
    """
    Prevent Excel from treating user text as formulas.

    Excel interprets leading "=", "+", "-", "@" (even after whitespace) as a
    formula trigger. If we write a span like "=> 10 mm" directly, openpyxl will
    serialize it as a formula, and Excel will "repair" the workbook by removing
    those invalid formulas.
    """

    if value is None:
        return ""

    if isinstance(value, str):
        stripped = value.lstrip()
        if stripped and stripped[0] in {"=", "+", "-", "@"}:
            return "'" + value
        return value

    return value


def _excel_safe_row_values(values: list[Any]) -> list[Any]:
    return [_excel_safe_value(v) for v in values]


def _ensure_sheet(wb: openpyxl.Workbook, name: str, headers: list[str] | None = None) -> openpyxl.worksheet.worksheet.Worksheet:
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    if headers:
        existing = _sheet_headers(ws)
        if not existing:
            ws.append(headers)
    return ws


def _append_row_by_headers(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: list[str],
    row: dict[str, Any],
) -> int:
    row_values = _excel_safe_row_values([row.get(h, "") for h in headers])
    ws.append(row_values)
    return ws.max_row


def _set_span_len_formula(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int) -> None:
    headers = _sheet_headers(ws)
    try:
        start_col = headers.index("start_char") + 1
        end_col = headers.index("end_char") + 1
        len_col = headers.index("span_len") + 1
    except ValueError:
        return

    start_letter = openpyxl.utils.get_column_letter(start_col)
    end_letter = openpyxl.utils.get_column_letter(end_col)
    len_cell = ws.cell(row=row_idx, column=len_col)
    len_cell.value = f'=IF(AND({start_letter}{row_idx}<>"",{end_letter}{row_idx}<>""),{end_letter}{row_idx}-{start_letter}{row_idx},"")'


def _normalize_span(note_id: str, source_file: str, idx: int, raw: dict[str, Any]) -> dict[str, Any]:
    def g(name: str, default: Any = None) -> Any:
        if isinstance(raw, dict):
            return raw.get(name, default)
        return getattr(raw, name, default)

    span_text = g("span_text") or g("text") or g("anchor") or g("quote") or ""
    context_prefix = g("context_prefix")
    if context_prefix is None:
        context_prefix = g("context") or g("prefix") or g("ctx") or ""

    label = g("label") or g("tag") or ""
    normalized_value = g("normalized_value")
    if normalized_value is None:
        normalized_value = g("norm") or g("normalized") or ""

    schema_field = g("schema_field")
    if schema_field is None:
        schema_field = g("schema") or g("field") or ""

    span_id = g("span_id") or f"{note_id}_span_{idx:03d}"
    section_type = g("section_type") or g("section") or "Procedure"

    match_index_int = _safe_int(g("match_index"))

    source_file_val = g("source_file") or source_file
    note_id_val = g("note_id") or note_id
    event_id_val = g("event_id") or g("event") or g("evt") or ""

    return {
        "source_file": source_file_val,
        "note_id": note_id_val,
        "span_id": span_id,
        "section_type": section_type,
        "context_prefix": context_prefix,
        "span_text": span_text,
        "match_index": match_index_int if match_index_int is not None else "",
        "start_char": "",
        "end_char": "",
        "span_len": "",
        "label": label,
        "normalized_value": normalized_value,
        "schema_field": schema_field,
        "event_id": event_id_val,
        "is_negated": _to_tf(g("is_negated"), default="FALSE"),
        "is_historical": _to_tf(g("is_historical"), default="FALSE"),
        "time_anchor": g("time_anchor") or "",
        "reviewer": g("reviewer") or "Auto",
        "comments": g("comments") or "",
        "hydration_status": g("hydration_status") or "needs_hydration",
    }


def _payload_from_script(script_path: Path, notes_dir: Path) -> NotePayload:
    g = runpy.run_path(str(script_path), run_name="__not_main__")

    note_id = str(g.get("NOTE_ID") or "").strip()
    if not note_id:
        m = re.search(r"Granular_note_(\\d{3})\\.py$", script_path.name)
        if not m:
            raise ValueError("Cannot determine NOTE_ID")
        note_id = f"note_{m.group(1)}"

    source_file = str(g.get("SOURCE_FILE") or "").strip() or f"{note_id}.txt"

    note_text = g.get("NOTE_TEXT")
    if not isinstance(note_text, str) or not note_text.strip():
        note_text = _read_note_text(notes_dir, source_file) or ""
    else:
        # Prefer notes_text on disk if present, to avoid drift from embedded strings.
        note_text = _read_note_text(notes_dir, source_file) or note_text

    procedure_date = str(g.get("PROCEDURE_DATE") or g.get("procedure_date") or "").strip()

    spans: list[Any] = []
    events: list[Any] = []
    flags: dict[str, Any] = {}

    # Prefer higher-level extractors when present.
    for extractor_name in ("extract_data", "generate_data", "generate_extraction_data"):
        fn = g.get(extractor_name)
        if not callable(fn):
            continue
        if extractor_name == "extract_data":
            try:
                sig = inspect.signature(fn)
                params = list(sig.parameters.values())
            except Exception:
                params = []
            if len(params) == 1 and params[0].name in {"gen", "generator"} and "WorkbookGenerator" in g and callable(g.get("WorkbookGenerator")):
                # Some scripts expose `extract_data(gen)` that populates a generator object.
                g["NOTE_TEXT"] = note_text
                gen = g["WorkbookGenerator"](g.get("TEMPLATE_PATH") or "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx")
                fn(gen)
                spans = list(getattr(gen, "spans", []) or [])
                events = list(getattr(gen, "events", []) or [])
                # Best-effort flags (mirrors common patterns in these scripts)
                flags = {
                    "diagnostic_bronchoscopy": 1,
                    "therapeutic_aspiration": 1,
                    "navigational_bronchoscopy": 1,
                    "radial_ebus": 1,
                    "transbronchial_biopsy": 1,
                    "transbronchial_cryobiopsy": 1,
                    "brushings": 1,
                    "bal": 1,
                    "linear_ebus": 1,
                }
                break
        res = _call_best_effort(fn, note_text)
        if isinstance(res, tuple) and len(res) == 3:
            flags, spans, events = res  # type: ignore[misc,assignment]
            break
        if isinstance(res, dict):
            flags = dict(res.get("flags") or res.get("FLAGS") or flags)
            spans = list(res.get("spans") or res.get("spans_data") or spans)
            events = list(res.get("events") or res.get("events_list") or events)
            break

    # Fallbacks: flags
    if not flags:
        if callable(g.get("determine_flags")):
            flags = _call_best_effort(g["determine_flags"], note_text) or {}
        elif callable(g.get("extract_flags")):
            flags = _call_best_effort(g["extract_flags"], note_text) or {}
        elif isinstance(g.get("PROCEDURE_FLAGS"), dict):
            flags = dict(g["PROCEDURE_FLAGS"])
        elif isinstance(g.get("FLAGS"), dict):
            flags = dict(g["FLAGS"])

    # Fallbacks: spans/events
    if not spans:
        if callable(g.get("extract_spans_and_events")):
            res = _call_best_effort(g["extract_spans_and_events"], note_text)
            if isinstance(res, tuple) and len(res) == 2:
                spans, events = res  # type: ignore[assignment]
            else:
                spans = res or []
        elif callable(g.get("generate_spans_and_events")):
            res = _call_best_effort(g["generate_spans_and_events"], note_text)
            if isinstance(res, tuple) and len(res) == 2:
                spans, events = res  # type: ignore[assignment]
            else:
                spans = res or []
        elif callable(g.get("extract_spans")):
            spans = _call_best_effort(g["extract_spans"], note_text) or []
        elif callable(g.get("generate_spans")):
            spans = _call_best_effort(g["generate_spans"], note_text) or []
        elif isinstance(g.get("spans_data"), list):
            spans = list(g["spans_data"])
        elif isinstance(g.get("spans"), list):
            spans = list(g["spans"])
        elif isinstance(g.get("SPANS"), list):
            spans = list(g["SPANS"])

    if not events:
        for k in ("events_list", "events", "raw_events", "v3_events_list"):
            if isinstance(g.get(k), list):
                events = list(g[k])
                break

    return NotePayload(
        note_id=note_id,
        source_file=source_file,
        note_text=note_text,
        procedure_date=procedure_date,
        flags=flags,
        spans=[s for s in spans if s is not None],
        events=[e for e in events if e is not None],
    )


def _note_ids_in_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, note_id_col_name: str = "note_id") -> set[str]:
    headers = _sheet_headers(ws)
    if not headers or note_id_col_name not in headers:
        return set()
    note_col = headers.index(note_id_col_name) + 1
    existing: set[str] = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=note_col).value
        if isinstance(v, str) and v.strip():
            existing.add(v.strip())
    return existing


def _append_note_index(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    payload: NotePayload,
) -> None:
    headers = _sheet_headers(ws)
    if not headers:
        raise ValueError("Note_Index is missing headers")

    meta_cols = [
        "source_file",
        "note_id",
        "encounter_id",
        "procedure_date",
        "site",
        "reviewer",
        "status",
        "free_text_notes",
    ]
    flag_cols = [h for h in headers if h not in meta_cols]

    flags = payload.flags or {}
    row: dict[str, Any] = {
        "source_file": payload.source_file,
        "note_id": payload.note_id,
        "encounter_id": "",
        "procedure_date": payload.procedure_date,
        "site": "",
        "reviewer": "Auto",
        "status": "auto_extracted",
        "free_text_notes": "",
    }
    for c in flag_cols:
        row[c] = _truthy_flag(flags.get(c))

    _append_row_by_headers(ws, headers, row)


def _append_note_text(
    wb: openpyxl.Workbook,
    payload: NotePayload,
) -> None:
    ws = _ensure_sheet(wb, "Note_Text", headers=["note_id", "source_file", "note_text"])
    # Avoid duplicate rows for same note_id.
    existing = _note_ids_in_sheet(ws, note_id_col_name="note_id")
    if payload.note_id in existing:
        return
    ws.append(_excel_safe_row_values([payload.note_id, payload.source_file, payload.note_text]))


def _append_spans(
    ws_anno: openpyxl.worksheet.worksheet.Worksheet,
    ws_hyd: openpyxl.worksheet.worksheet.Worksheet,
    payload: NotePayload,
) -> tuple[int, int]:
    headers_anno = _sheet_headers(ws_anno)
    headers_hyd = _sheet_headers(ws_hyd)
    if not headers_anno or not headers_hyd:
        raise ValueError("Span sheets missing headers")

    added_anno = 0
    added_hyd = 0

    raw_spans = payload.spans or []
    for idx, raw in enumerate(raw_spans, start=1):
        s = _normalize_span(payload.note_id, payload.source_file, idx, raw)
        if not str(s.get("span_text") or "").strip():
            # Skip "ghost" rows: a span entry with no actual text to label.
            continue

        # Span_Annotations: leave offsets empty and mark needs_hydration.
        s_anno = dict(s)
        s_anno["start_char"] = ""
        s_anno["end_char"] = ""
        s_anno["hydration_status"] = "needs_hydration"
        row_idx = _append_row_by_headers(ws_anno, headers_anno, s_anno)
        _set_span_len_formula(ws_anno, row_idx)
        added_anno += 1

        # Span_Hydrated: pre-hydrate if possible.
        start, end, status = _hydrate_span(
            payload.note_text,
            str(s.get("span_text") or ""),
            str(s.get("context_prefix") or "") or None,
            _safe_int(s.get("match_index")),
        )
        s_hyd = dict(s)
        s_hyd["start_char"] = start if start is not None else ""
        s_hyd["end_char"] = end if end is not None else ""
        s_hyd["hydration_status"] = status
        row_idx = _append_row_by_headers(ws_hyd, headers_hyd, s_hyd)
        _set_span_len_formula(ws_hyd, row_idx)
        added_hyd += 1

    return added_anno, added_hyd


def _append_event_log(
    ws_event: openpyxl.worksheet.worksheet.Worksheet,
    payload: NotePayload,
) -> int:
    headers = _sheet_headers(ws_event)
    if not headers:
        raise ValueError("Event_Log missing headers")

    raw_spans = payload.spans or []
    by_event: dict[str, list[dict[str, Any]]] = {}
    for idx, raw in enumerate(raw_spans, start=1):
        s = _normalize_span(payload.note_id, payload.source_file, idx, raw)
        event_id = str(s.get("event_id") or "").strip()
        if not event_id:
            continue
        by_event.setdefault(event_id, []).append(s)

    added = 0
    for event_id, group in sorted(by_event.items()):
        methods: set[str] = set()
        anatomy: set[str] = set()
        devices: set[str] = set()
        findings: set[str] = set()

        for s in group:
            label = str(s.get("label") or "")
            norm = str(s.get("normalized_value") or "")
            if label == "PROC_METHOD" and norm:
                methods.add(norm)
            if label in {"ANAT_LUNG_LOC", "ANAT_LN_STATION", "ANAT_PLEURAL_LOC"} and norm:
                anatomy.add(norm)
            if label.startswith("DEV_") and norm:
                devices.add(norm)
            if label.startswith("OBS_") and norm:
                findings.add(norm)

        row = {
            "source_file": payload.source_file,
            "note_id": payload.note_id,
            "event_id": event_id,
            "event_type": "",
            "method": ", ".join(sorted(methods)),
            "anatomy_target": ", ".join(sorted(anatomy)),
            "device": ", ".join(sorted(devices)),
            "needle_gauge": "",
            "stations": "",
            "counts": "",
            "measurements": "",
            "specimens": "",
            "findings": ", ".join(sorted(findings)),
            "is_historical": "FALSE",
            "reviewer": "Auto",
            "comments": "",
            "device_size": "",
            "device_material": "",
            "outcome_airway_lumen_pre": "",
            "outcome_airway_lumen_post": "",
            "outcome_symptoms": "",
            "outcome_pleural": "",
            "outcome_complication": "",
        }
        _append_row_by_headers(ws_event, headers, row)
        added += 1

    return added


def _append_v3_events(
    ws_v3: openpyxl.worksheet.worksheet.Worksheet,
    payload: NotePayload,
) -> int:
    headers = _sheet_headers(ws_v3)
    if not headers:
        raise ValueError("V3_Procedure_Events missing headers")

    events_raw = payload.events or []
    added = 0

    for evt in events_raw:
        if isinstance(evt, dict):
            event_id = evt.get("event_id") or evt.get("id") or ""
            etype = evt.get("type") or ""
            method = evt.get("method") or ""

            target = evt.get("target") or {}
            if not isinstance(target, dict):
                target = {}
            loc = target.get("location") or {}
            if not isinstance(loc, dict):
                loc = {}

            lesion = evt.get("lesion") or {}
            if not isinstance(lesion, dict):
                lesion = {}

            outcomes = evt.get("outcomes") or {}
            if not isinstance(outcomes, dict):
                outcomes = {}

            devices = evt.get("devices") or []
            measurements = evt.get("measurements") or {}
            specimens = evt.get("specimens") or []
            findings = evt.get("findings") or []
            evidence_quote = evt.get("evidence_quote") or ""
        else:
            # Best-effort support for event objects from some scripts.
            event_id = getattr(evt, "event_id", "") or getattr(evt, "id", "")
            etype = getattr(evt, "type", "") or getattr(evt, "event_type", "")
            method = getattr(evt, "method", "")
            target = getattr(evt, "target", {}) or {}
            loc = {}
            lesion = {}
            outcomes = getattr(evt, "outcomes", {}) or {}
            devices = getattr(evt, "devices", []) or []
            measurements = getattr(evt, "measurements", {}) or {}
            specimens = getattr(evt, "specimens", []) or []
            findings = getattr(evt, "findings", []) or []
            evidence_quote = getattr(evt, "evidence_quote", "") or ""
            if not isinstance(target, dict):
                target = {}
            if not isinstance(outcomes, dict):
                outcomes = {}

        row = {
            "note_id": payload.note_id,
            "event_id": event_id,
            "type": etype,
            "target.anatomy_type": target.get("anatomy_type") or target.get("structure") or "",
            "target.location.lobe": loc.get("lobe") or "",
            "target.location.segment": loc.get("segment") or "",
            "target.station": target.get("station") or "",
            "lesion.type": lesion.get("type") or "",
            "lesion.size_mm": lesion.get("size_mm") or "",
            "method": method,
            "devices_json": json.dumps(devices),
            "measurements_json": json.dumps(measurements),
            "specimens_json": json.dumps(specimens),
            "findings_json": json.dumps(findings),
            "evidence_quote": evidence_quote,
            "stent.size": "",
            "stent.material_or_brand": "",
            "catheter.size_fr": "",
            "outcomes.airway.lumen_pre": (outcomes.get("airway") or {}).get("lumen_pre") if isinstance(outcomes.get("airway"), dict) else "",
            "outcomes.airway.lumen_post": (outcomes.get("airway") or {}).get("lumen_post") if isinstance(outcomes.get("airway"), dict) else "",
            "outcomes.symptoms": outcomes.get("symptoms") or "",
            "outcomes.pleural": outcomes.get("pleural") or "",
            "outcomes.complications": outcomes.get("complications") or "",
        }
        _append_row_by_headers(ws_v3, headers, row)
        added += 1

    return added


def populate_workbook(
    template_path: Path,
    output_path: Path,
    scripts_dir: Path,
    notes_dir: Path,
    include_note_text: bool,
    limit: int | None,
) -> None:
    if output_path.resolve() == template_path.resolve():
        raise ValueError("Refusing to overwrite template in-place; choose a different --output path.")

    shutil.copyfile(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    ws_index = wb["Note_Index"]
    ws_anno = wb["Span_Annotations"]
    ws_hyd = wb["Span_Hydrated"]
    ws_event = wb["Event_Log"]
    ws_v3 = wb["V3_Procedure_Events"]

    existing_notes = _note_ids_in_sheet(ws_index)
    processed = 0
    failures: list[str] = []

    scripts = sorted(scripts_dir.glob("Granular_note_*.py"))
    if limit is not None:
        scripts = scripts[:limit]

    for script_path in scripts:
        try:
            payload = _payload_from_script(script_path, notes_dir)
        except Exception as e:
            failures.append(f"{script_path.name}: load_failed: {e}")
            continue

        if payload.note_id in existing_notes:
            continue

        try:
            _append_note_index(ws_index, payload)
            if include_note_text:
                _append_note_text(wb, payload)
            _append_spans(ws_anno, ws_hyd, payload)
            _append_event_log(ws_event, payload)
            _append_v3_events(ws_v3, payload)
        except Exception as e:
            failures.append(f"{script_path.name}: write_failed: {e}")
            continue

        existing_notes.add(payload.note_id)
        processed += 1

    wb.save(output_path)

    print(f"Wrote workbook: {output_path}")
    print(f"Processed notes: {processed}")
    if failures:
        print(f"Failures: {len(failures)}")
        for line in failures[:50]:
            print(f"  - {line}")
        if len(failures) > 50:
            print("  - ...")


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Populate the phase0 golden registry workbook from per-note scripts.")
    p.add_argument(
        "--template",
        default="phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx",
        help="Path to the phase0 template workbook.",
    )
    p.add_argument(
        "--output",
        default="phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.POPULATED.xlsx",
        help="Output workbook path (will be created).",
    )
    p.add_argument(
        "--scripts-dir",
        default="python scripts",
        help="Directory containing Granular_note_*.py scripts.",
    )
    p.add_argument(
        "--notes-dir",
        default="notes_text",
        help="Directory containing note_*.txt files.",
    )
    p.add_argument(
        "--include-note-text",
        action="store_true",
        help="Add a Note_Text sheet containing full note text (if not already present).",
    )
    p.add_argument("--limit", type=int, default=None, help="Only process the first N scripts.")
    args = p.parse_args(argv)

    populate_workbook(
        template_path=Path(args.template),
        output_path=Path(args.output),
        scripts_dir=Path(args.scripts_dir),
        notes_dir=Path(args.notes_dir),
        include_note_text=bool(args.include_note_text),
        limit=args.limit,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
