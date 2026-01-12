import json
import os
import re

import openpyxl

# =============================================================================
# INPUT DATA
# =============================================================================
NOTE_ID = "note_043"
SOURCE_FILE = "note_043.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"
PROCEDURE_DATE = "2025-12-22"

NOTE_TEXT = ""
NOTE_PATH = os.path.join("notes_text", SOURCE_FILE)
if os.path.exists(NOTE_PATH):
    NOTE_TEXT = open(NOTE_PATH, "r", encoding="utf-8").read()


# =============================================================================
# FLAGS
# =============================================================================

PROCEDURE_FLAGS_LIST = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]


def determine_flags(text: str) -> dict[str, int]:
    flags = {k: 0 for k in PROCEDURE_FLAGS_LIST}
    tl = text.lower()

    # Pleural fibrinolysis via chest tube
    if "fibrinolysis" in tl or "tpa" in tl or "dnase" in tl or "32562" in tl:
        flags["fibrinolytic_therapy"] = 1
    if "chest tube insertion" in tl or "via chest tube" in tl or "unclamp chest tube" in tl:
        flags["chest_tube"] = 1
    return flags


# =============================================================================
# SPANS
# =============================================================================

spans: list[dict[str, object]] = []


def add_span(
    span_text: str,
    label: str,
    normalized_value: str,
    schema_field: str,
    event_id: str,
    context_prefix: str | None = None,
    match_index: int | None = None,
) -> None:
    if not NOTE_TEXT:
        return
    if span_text not in NOTE_TEXT:
        return
    spans.append(
        {
            "span_text": span_text,
            "label": label,
            "normalized_value": normalized_value,
            "schema_field": schema_field,
            "event_id": event_id,
            "context_prefix": context_prefix or "",
            "match_index": match_index if match_index is not None else "",
            "is_negated": "FALSE",
            "is_historical": "FALSE",
            "comments": "",
        }
    )


EVT_US = "evt_01_us"
EVT_FIB = "evt_02_fibrinolysis"
EVT_OUT = "evt_03_outcome"

add_span("Ultrasound, chest", "PROC_METHOD", "chest_ultrasound", "method", EVT_US)
add_span("Pleural Effusion:", "OBS_LESION", "pleural_effusion", "findings_json", EVT_US)
add_span("Date of chest tube insertion: 12/22/2025", "TIME_ANCHOR", "2025-12-22", "procedure_date", EVT_FIB)
add_span("Side: left", "ANAT_PLEURAL_LOC", "left", "target.side", EVT_FIB)
add_span("5 mg/5 mg tPA/Dnasedose #: 2", "MEDICATION", "tpa_dnase_dose_2", "medications", EVT_FIB)
add_span("Unclamp chest tube in 1 hour", "PROC_ACTION", "unclamp_in_1h", "plan", EVT_FIB)
add_span("There were no immediate complications.", "OUTCOME_COMPLICATION", "none", "outcomes.complications", EVT_OUT)


# =============================================================================
# HYDRATION
# =============================================================================


def hydrate_span(text: str, span_text: str, context_prefix: str | None = None, match_index: int | None = None):
    matches = [m.start() for m in re.finditer(re.escape(span_text), text)]
    if not matches:
        return "", "", "not_found"
    if len(matches) == 1:
        start = matches[0]
        return start, start + len(span_text), "hydrated_unique"
    if context_prefix:
        for start in matches:
            window = text[max(0, start - 120):start]
            if context_prefix in window:
                return start, start + len(span_text), "hydrated_prefix_window"
    if match_index is not None and 0 <= match_index < len(matches):
        start = matches[match_index]
        return start, start + len(span_text), "hydrated_match_index"
    return "", "", f"ambiguous_count={len(matches)}"


# =============================================================================
# WORKBOOK GENERATION
# =============================================================================


def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    wb = openpyxl.load_workbook(template_path)

    # Note_Text (optional)
    if "Note_Text" not in wb.sheetnames:
        wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # Note_Index
    ws_index = wb["Note_Index"]
    flags = determine_flags(NOTE_TEXT)
    row_meta = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "auto_extracted", ""]
    row_flags = [flags.get(k, 0) for k in PROCEDURE_FLAGS_LIST]
    ws_index.append(row_meta + row_flags)

    # Span_Annotations & Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hyd = wb["Span_Hydrated"]
    for i, s in enumerate(spans, start=1):
        span_id = f"{NOTE_ID}_s{i:03d}"
        start, end, status = hydrate_span(NOTE_TEXT, str(s["span_text"]), str(s.get("context_prefix") or "") or None, None)
        base = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", s.get("context_prefix", ""), s["span_text"], s.get("match_index", ""),
            "", "", "", s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s.get("is_negated", "FALSE"), s.get("is_historical", "FALSE"), "", "Auto", s.get("comments", ""), "needs_hydration",
        ]
        ws_anno.append(base)
        hyd = list(base)
        hyd[7] = start
        hyd[8] = end
        hyd[19] = status
        ws_hyd.append(hyd)

    # Event_Log (minimal)
    ws_event = wb["Event_Log"]
    ws_event.append([SOURCE_FILE, NOTE_ID, EVT_US, "Procedure", "Chest Ultrasound", "Pleura", "", "", "", "", "", "", "Pleural effusion", "FALSE", "Auto", "", "", "", "", "", "", "", ""])
    ws_event.append([SOURCE_FILE, NOTE_ID, EVT_FIB, "Procedure", "Fibrinolytic therapy", "Pleura", "", "", "", "", "tPA/DNase dose #2", "", "", "FALSE", "Auto", "", "", "", "", "", "", "", ""])
    ws_event.append([SOURCE_FILE, NOTE_ID, EVT_OUT, "Outcome", "", "", "", "", "", "", "", "", "", "FALSE", "Auto", "", "", "", "", "", "", "", "None"])

    wb.save(output_path)
    print(f"Generated {output_path}")


if __name__ == "__main__":
    generate_excel()

